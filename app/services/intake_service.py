"""Universal intake orchestration service."""
from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import os
import re
import uuid
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.core.config import settings
from app.models.intake import IntakeSession, IntakeItem
from app.models.analysis import AnalysisResult
from app.models.bom import BOM
from app.models.user import User
from app.schemas.intake import (
    IntakeInputType,
    IntakeIntent,
    IntakeParseRequest,
    IntakeSubmitRequest,
)
from app.services import (
    analyzer_service,
    bom_service,
    pricing_service,
    project_service,
    resolver_service,
    review_service,
    vendor_service,
)
from app.services.strategy_service import build_strategy_output
from app.services.procurement_planner import generate_procurement_plan

logger = logging.getLogger("intake_service")

CATEGORY_KEYWORDS = {
    "fastener": ["bolt", "nut", "screw", "washer", "rivet", "fastener", "stud", "threaded"],
    "electrical": ["wire", "cable", "connector", "terminal", "relay", "switch", "sensor", "harness"],
    "electronics": ["resistor", "capacitor", "inductor", "ic", "microcontroller", "pcb", "chip", "led", "diode", "transistor", "smt"],
    "mechanical": ["bracket", "housing", "shaft", "gear", "spacer", "plate", "frame", "assembly"],
    "raw_material": ["aluminum", "steel", "copper", "brass", "titanium", "nylon", "abs", "polycarbonate", "stainless", "sheet", "bar", "rod", "tube"],
    "sheet_metal": ["sheet metal", "laser cut", "bend", "bent", "formed", "stamped"],
    "machined": ["machined", "cnc", "turned", "milled", "drilled", "threaded"],
    "custom_mechanical": ["custom", "fabricated", "prototype"],
    "pneumatic": ["pneumatic", "air valve", "air cylinder", "fitting", "hose"],
    "hydraulic": ["hydraulic", "seal", "pump", "valve", "hose"],
    "optical": ["lens", "optic", "optical", "camera", "fiber"],
    "thermal": ["heater", "thermal", "heat sink", "cooling", "fan", "radiator"],
    "cable_wiring": ["cable", "harness", "loom", "wire"],
}

PROCESS_KEYWORDS = {
    "machining": ["cnc", "machined", "milled", "turned", "drilled"],
    "sheet_metal": ["laser cut", "bent", "sheet metal", "stamped", "formed"],
    "injection_molding": ["injection molded", "injection molding", "molded"],
    "casting": ["cast", "casting"],
    "forging": ["forged", "forging"],
    "welding": ["weld", "welded", "welding"],
    "pcb_assembly": ["pcb assembly", "smt", "soldered", "reflow"],
    "cable_assembly": ["harness", "loom", "cable assembly"],
    "extrusion": ["extruded", "extrusion"],
}

INTENT_KEYWORDS = {
    "deep_search": ["deep search", "search deeply", "find supplier", "find suppliers", "source suppliers"],
    "rfq": ["rfq", "request for quote", "request quotes", "send quote"],
    "compare": ["compare", "comparison", "lowest cost", "best lead time"],
    "price_check": ["price", "cost", "costing", "estimate", "benchmark"],
    "vendor_search": ["vendor", "supplier", "match suppliers", "supplier match"],
    "research_product": ["research product", "product research", "find product"],
}

MATERIAL_HINTS = [
    "aluminum", "steel", "copper", "brass", "titanium", "nylon", "abs", "pc", "polycarbonate",
    "stainless", "rubber", "silicone", "glass", "ceramic", "carbon fiber", "sheet", "bar", "rod"
]

UNIT_PATTERNS = [
    (r"\b(?:pcs?|pieces?|units?|ea|each)\b", "pcs"),
    (r"\b(?:kg|kilograms?)\b", "kg"),
    (r"\b(?:g|grams?)\b", "g"),
    (r"\b(?:mm|millimeters?)\b", "mm"),
    (r"\b(?:cm|centimeters?)\b", "cm"),
    (r"\b(?:m|meters?)\b", "m"),
    (r"\b(?:l|liters?)\b", "l"),
    (r"\b(?:ml|milliliters?)\b", "ml"),
    (r"\b(?:reels?|rolls?)\b", "reel"),
    (r"\b(?:sets?)\b", "set"),
]

SPEC_PATTERNS = {
    "thickness": [
        r"(?:thickness|thk|t)\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(mm|cm|in|inch|inches|\")",
    ],
    "diameter": [
        r"(?:diameter|dia|ø)\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(mm|cm|in|inch|inches|\")",
    ],
    "length": [
        r"(?:length|len)\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(mm|cm|m|in|inch|inches|\")",
    ],
    "width": [
        r"(?:width|w)\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(mm|cm|m|in|inch|inches|\")",
    ],
    "height": [
        r"(?:height|h)\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(mm|cm|m|in|inch|inches|\")",
    ],
    "voltage": [
        r"(\d+(?:\.\d+)?)\s*(v|vac|vdc)\b",
    ],
    "current": [
        r"(\d+(?:\.\d+)?)\s*(a|amp|amps|ma)\b",
    ],
    "power": [
        r"(\d+(?:\.\d+)?)\s*(w|kw|mw)\b",
    ],
    "resistance": [
        r"(\d+(?:\.\d+)?)\s*(ohm|ohms|kω|mω|kohm|mohm)\b",
    ],
    "capacitance": [
        r"(\d+(?:\.\d+)?)\s*(uf|nf|pf|farad|farads)\b",
    ],
    "tolerance": [
        r"([±]?\d+(?:\.\d+)?)\s*%",
    ],
    "finish": [
        r"\b(anodized|plated|galvanized|painted|polished|coated|passivated)\b",
    ],
    "grade": [
        r"\b(a2|a4|316|304|6061|7075|fr4|ul94-v0|nema)\b",
    ],
    "color": [
        r"\b(black|white|red|green|blue|yellow|gray|silver|gold|natural)\b",
    ],
}

QUANTITY_PATTERNS = [
    r"\b(\d+(?:\.\d+)?)\s*(pcs?|pieces?|units?|ea|each|kg|g|mm|cm|m|l|ml|reels?|rolls?|sets?)\b",
    r"\bqty\s*[:=]?\s*(\d+(?:\.\d+)?)\b",
    r"^\s*(\d+(?:\.\d+)?)\s*[xX]\s*",
    r"\b(\d+(?:\.\d+)?)\s*[xX]\s*(?!\s*\d)",
]

SOURCING_FALLBACKS = {
    "buyer_action": "Add quantity, material grade, or drawing for stronger matching.",
    "supplier_action": "Use vendor discovery to rank suppliers by capability and lead time.",
    "rfq_action": "Generate an RFQ once quantity/specs are sufficient.",
}


def _stable_hash(value: Any) -> str:
    encoded = json.dumps(value, sort_keys=True, default=str, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _safe_lower(value: Optional[str]) -> str:
    return (value or "").strip().lower()


def _detect_input_type(
    *,
    explicit_type: Optional[str],
    raw_input_text: str,
    file_name: Optional[str],
    voice_transcript: Optional[str],
) -> str:
    if explicit_type and explicit_type != "auto":
        return explicit_type
    if voice_transcript:
        return "voice"
    if file_name:
        lower = file_name.lower()
        if lower.endswith((".csv", ".xlsx", ".xls", ".tsv")):
            return "bom"
        if lower.endswith((".txt", ".md")):
            return "free_text"
        return "file"
    text = raw_input_text or ""
    if "\n" in text and len(text.splitlines()) > 1:
        return "bom"
    if any(sep in text for sep in ("|", "\t")):
        return "bom"
    if any(token in text.lower() for token in ("material", "component", "item", "part", "qty", "quantity")):
        return "item"
    return "free_text"


def _detect_intent(explicit_intent: Optional[str], text: str) -> str:
    if explicit_intent and explicit_intent != "auto":
        return explicit_intent
    t = _safe_lower(text)
    for intent, keywords in INTENT_KEYWORDS.items():
        if any(k in t for k in keywords):
            return intent
    return "source"


def _normalize_unit(text: str) -> Optional[str]:
    t = _safe_lower(text)
    for pattern, unit in UNIT_PATTERNS:
        if re.search(pattern, t, flags=re.I):
            return unit
    return None


def _detect_category(text: str) -> str:
    t = _safe_lower(text)
    for category, keywords in CATEGORY_KEYWORDS.items():
        if any(k in t for k in keywords):
            return category
    return "standard"


def _detect_process(text: str) -> Optional[str]:
    t = _safe_lower(text)
    for process, keywords in PROCESS_KEYWORDS.items():
        if any(k in t for k in keywords):
            return process
    return None


def _detect_material(text: str) -> Optional[str]:
    t = _safe_lower(text)
    for hint in MATERIAL_HINTS:
        if hint in t:
            return hint
    return None


def _extract_quantity(text: str) -> Tuple[float, Optional[str]]:
    t = text.strip()
    for pattern in QUANTITY_PATTERNS:
        m = re.search(pattern, t, flags=re.I)
        if m:
            raw = m.group(1)
            try:
                qty = float(raw)
                unit = _normalize_unit(t)
                return qty, unit
            except ValueError:
                continue
    return 1.0, _normalize_unit(t)


def _extract_specs(text: str) -> Dict[str, Any]:
    specs: Dict[str, Any] = {}
    t = text.strip()

    for key, patterns in SPEC_PATTERNS.items():
        for pattern in patterns:
            m = re.search(pattern, t, flags=re.I)
            if m:
                if key in {"thickness", "diameter", "length", "width", "height", "voltage", "current", "power", "resistance", "capacitance", "tolerance"}:
                    if len(m.groups()) >= 2:
                        specs[key] = {"value": m.group(1), "unit": m.group(2)}
                    else:
                        specs[key] = m.group(1)
                else:
                    specs[key] = m.group(1)
                break

    if "material" not in specs:
        material = _detect_material(text)
        if material:
            specs["material_hint"] = material

    return specs


def _extract_item_name(text: str) -> str:
    t = text.strip()
    if not t:
        return ""
    t = re.sub(r"\bqty\s*[:=]?\s*\d+(?:\.\d+)?\b", "", t, flags=re.I)
    t = re.sub(r"^\s*\d+(?:\.\d+)?\s*[xX]\s*", "", t)
    t = re.sub(r"\b\d+(?:\.\d+)?\s*(pcs?|pieces?|units?|ea|each|kg|g|mm|cm|m|l|ml|reels?|rolls?|sets?)\b", "", t, flags=re.I)
    t = re.sub(r"[\|\t;]+", " ", t)
    t = re.sub(r"\s{2,}", " ", t).strip()
    return t[:200]


def _missing_data_warnings(item: Dict[str, Any]) -> List[str]:
    warnings: List[str] = []
    if not item.get("quantity") or float(item.get("quantity") or 0) <= 0:
        warnings.append("Quantity missing or invalid.")
    if not item.get("material") and item.get("category") in {"raw_material", "machined", "sheet_metal", "custom_mechanical"}:
        warnings.append("Material or grade would improve sourcing accuracy.")
    if item.get("category") in {"electronics", "electrical"} and not item.get("specs"):
        warnings.append("Electrical/electronics spec details are sparse.")
    if item.get("category") in {"custom_mechanical", "machined", "sheet_metal"} and not item.get("process"):
        warnings.append("Process information is missing for custom sourcing.")
    return warnings


def _score_confidence(
    *,
    item_name: str,
    category: str,
    material: Optional[str],
    process: Optional[str],
    specs: Dict[str, Any],
    quantity: float,
) -> float:
    score = 0.35
    if item_name:
        score += 0.15
    if category and category != "standard":
        score += 0.1
    if material:
        score += 0.15
    if process:
        score += 0.1
    if specs:
        score += min(0.15, 0.03 * len(specs))
    if quantity and quantity > 0:
        score += 0.05
    return round(min(score, 0.99), 2)


def _split_candidate_lines(text: str) -> List[str]:
    cleaned = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    lines = [line.strip() for line in cleaned.split("\n") if line.strip()]
    if len(lines) == 1:
        line = lines[0]
        if any(sep in line for sep in ("|", ";", "\t")):
            parts = [p.strip() for p in re.split(r"[|\t;]", line) if p.strip()]
            return parts or lines
    return lines


def _parse_candidate(text: str, line_no: int) -> Dict[str, Any]:
    quantity, unit = _extract_quantity(text)
    category = _detect_category(text)
    process = _detect_process(text)
    material = _detect_material(text)
    specs = _extract_specs(text)
    item_name = _extract_item_name(text)
    warnings = _missing_data_warnings(
        {
            "quantity": quantity,
            "material": material,
            "category": category,
            "process": process,
            "specs": specs,
        }
    )
    confidence = _score_confidence(
        item_name=item_name,
        category=category,
        material=material,
        process=process,
        specs=specs,
        quantity=quantity,
    )
    return {
        "line_no": line_no,
        "raw_text": text,
        "item_name": item_name,
        "category": category,
        "material": material,
        "process": process,
        "quantity": quantity,
        "unit": unit,
        "specs": specs,
        "confidence": confidence,
        "warnings": warnings,
    }


def _parse_text_to_items(text: str) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    candidates = _split_candidate_lines(text or "")
    parsed: List[Dict[str, Any]] = []
    for idx, candidate in enumerate(candidates, start=1):
        parsed.append(_parse_candidate(candidate, idx))

    if not parsed and text.strip():
        parsed.append(_parse_candidate(text.strip(), 1))

    categories = Counter([item["category"] for item in parsed])
    warnings = sorted({w for item in parsed for w in item.get("warnings", [])})
    suggestions: List[str] = []
    if any(item["category"] in {"custom_mechanical", "machined", "sheet_metal"} and not item["process"] for item in parsed):
        suggestions.append("Add process details such as machining, laser cut, bending, or welding.")
    if any(not item["material"] for item in parsed):
        suggestions.append("Add material or grade to improve vendor matching.")
    if any(item["quantity"] == 1.0 and not re.search(r"\bqty\b|\bpcs?\b|\bunits?\b", item["raw_text"], flags=re.I) for item in parsed):
        suggestions.append("Specify quantity to improve costing and RFQ generation.")
    summary = {
        "line_count": len(parsed),
        "category_breakdown": dict(categories),
        "warnings": warnings,
        "suggestions": suggestions,
    }
    return parsed, summary


def _read_csv_bytes(file_bytes: bytes) -> str:
    decoded = file_bytes.decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(decoded))
    rows = []
    for row in reader:
        if not row:
            continue
        rows.append(" | ".join(cell.strip() for cell in row if cell and cell.strip()))
    return "\n".join(rows)


def _read_excel_bytes(file_bytes: bytes) -> str:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows: List[str] = []
    for row in sheet.iter_rows(values_only=True):
        values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
        if values:
            rows.append(" | ".join(values))
    return "\n".join(rows)


def _read_file_as_text(file_bytes: bytes, filename: str) -> str:
    lower = (filename or "").lower()
    if lower.endswith(".csv") or lower.endswith(".tsv"):
        return _read_csv_bytes(file_bytes)
    if lower.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        return _read_excel_bytes(file_bytes)
    return file_bytes.decode("utf-8", errors="ignore")


def _normalize_request_text(
    *,
    raw_input_text: Optional[str],
    input_type: str,
    file_text: Optional[str],
    voice_transcript: Optional[str],
) -> Tuple[str, List[Dict[str, Any]], Dict[str, Any]]:
    pieces: List[str] = []
    if raw_input_text:
        pieces.append(raw_input_text.strip())
    if voice_transcript:
        pieces.append(voice_transcript.strip())
    if file_text:
        pieces.append(file_text.strip())

    combined = "\n".join([p for p in pieces if p])
    if not combined:
        combined = raw_input_text or voice_transcript or file_text or ""

    parsed, summary = _parse_text_to_items(combined)
    return combined, parsed, summary


def _build_synthetic_analyzer_output(
    *,
    session: IntakeSession,
    parsed_items: List[Dict[str, Any]],
    combined_text: str,
    summary: Dict[str, Any],
    delivery_location: str,
    target_currency: str,
    priority: str,
) -> Dict[str, Any]:
    components = []
    for item in parsed_items:
        components.append(
            {
                "item_id": f"INTAKE-{item['line_no']}",
                "description": item["item_name"] or item["raw_text"],
                "quantity": item["quantity"],
                "material": item["material"] or "",
                "mpn": "",
                "manufacturer": "",
                "notes": item["raw_text"],
                "category": item["category"],
                "classification_confidence": item["confidence"],
                "geometry": None,
                "tolerance": item["specs"].get("tolerance"),
                "material_form": item["specs"].get("material_hint") or item["material"],
                "secondary_ops": [item["process"]] if item["process"] else [],
                "specs": item["specs"],
                "procurement_class": "rfq_required" if item["category"] in {"custom_mechanical", "machined", "sheet_metal", "electronics"} else "catalog_purchase",
                "rfq_required": item["category"] in {"custom_mechanical", "machined", "sheet_metal", "electronics"},
                "drawing_required": item["category"] in {"custom_mechanical", "machined", "sheet_metal"},
                "is_custom": item["category"] in {"custom_mechanical", "machined", "sheet_metal"},
                "part_type": item["category"],
                "source_intake_type": session.input_type,
            }
        )

    category_counts = Counter([item["category"] for item in parsed_items])
    estimated_lines = len(parsed_items)
    estimated_total_qty = round(sum(float(item.get("quantity") or 1) for item in parsed_items), 2)

    return {
        "_meta": {
            "version": "universal-intake-v1",
            "generated_at": datetime.utcnow().isoformat(),
            "source": "intake_service.synthetic_analyzer",
            "input_type": session.input_type,
            "intent": session.intent,
            "delivery_location": delivery_location,
            "target_currency": target_currency,
            "priority": priority,
            "confidence_score": session.confidence_score,
        },
        "original_input": combined_text,
        "components": components,
        "summary": {
            "line_count": estimated_lines,
            "total_quantity": estimated_total_qty,
            "category_breakdown": dict(category_counts),
            "warnings": summary.get("warnings", []),
            "suggestions": summary.get("suggestions", []),
        },
        "_v2_full_report": {
            "section_1_executive_summary": {
                "input_type": session.input_type,
                "intent": session.intent,
                "summary": f"Parsed {estimated_lines} sourcing item(s) from universal intake.",
            },
            "section_2_component_breakdown": components,
            "section_3_sourcing_strategy": {
                "preferred_region": delivery_location,
                "target_currency": target_currency,
                "priority": priority,
                "recommended_action": session.intent,
            },
            "section_4_financial_summary": {
                "estimated_quantity": estimated_total_qty,
                "currency": target_currency,
            },
            "section_5_recommendation": {
                "next_step": "Create project and vendor shortlist",
            },
            "section_6_learning_snapshot": {
                "input_signature": _stable_hash({"text": combined_text, "items": components}),
            },
        },
    }


def _unique_suggestions(values: Iterable[str]) -> List[str]:
    seen = set()
    out = []
    for v in values:
        if not v:
            continue
        if v not in seen:
            seen.add(v)
            out.append(v)
    return out


def _resolve_session_identity(
    *,
    db: Session,
    namespace: str,
    idempotency_key: Optional[str],
    request_hash: str,
    user: Optional[User],
    session_token: Optional[str],
    input_type: str,
    intent: str,
    delivery_location: str,
    target_currency: str,
    priority: str,
    raw_input_text: Optional[str],
    normalized_text: Optional[str],
    voice_transcript: Optional[str],
    source_channel: str,
    metadata: Dict[str, Any],
    source_file_name: Optional[str] = None,
    source_file_type: Optional[str] = None,
    source_file_size: Optional[int] = None,
    source_file_path: Optional[str] = None,
    audio_file_name: Optional[str] = None,
    audio_file_type: Optional[str] = None,
    audio_file_size: Optional[int] = None,
    audio_file_path: Optional[str] = None,
) -> IntakeSession:
    idempotency_key = idempotency_key or request_hash[:24]
    existing = (
        db.query(IntakeSession)
        .filter(
            IntakeSession.namespace == namespace,
            IntakeSession.idempotency_key == idempotency_key,
        )
        .first()
    )
    if existing:
        if existing.request_hash and existing.request_hash != request_hash:
            raise ValueError("Idempotency key reuse detected with a different intake payload.")
        return existing

    session = IntakeSession(
        namespace=namespace,
        idempotency_key=idempotency_key,
        request_hash=request_hash,
        user_id=user.id if user else None,
        guest_session_id=None,
        session_token=session_token,
        input_type=input_type,
        intent=intent,
        source_channel=source_channel,
        raw_input_text=raw_input_text,
        normalized_text=normalized_text,
        voice_transcript=voice_transcript,
        source_file_name=source_file_name,
        source_file_type=source_file_type,
        source_file_size=source_file_size,
        source_file_path=source_file_path,
        audio_file_name=audio_file_name,
        audio_file_type=audio_file_type,
        audio_file_size=audio_file_size,
        audio_file_path=audio_file_path,
        delivery_location=delivery_location,
        target_currency=target_currency,
        priority=priority,
        status="received",
        parse_status="pending",
        analysis_status="pending",
        workflow_status="received",
        metadata_json=metadata or {},
        warnings=[],
        suggestions=[],
        parsed_payload={},
        normalized_payload={},
        analysis_payload={},
        preview_payload={},
    )
    db.add(session)
    db.flush()
    return session


def _serialize_item(item: IntakeItem) -> Dict[str, Any]:
    return {
        "line_no": item.line_no,
        "raw_text": item.raw_text,
        "item_name": item.item_name,
        "category": item.category,
        "material": item.material,
        "process": item.process,
        "quantity": item.quantity,
        "unit": item.unit,
        "specs": item.specs or {},
        "confidence": item.confidence,
        "warnings": item.warnings or [],
    }


def serialize_session(session: IntakeSession) -> Dict[str, Any]:
    return {
        "id": session.id,
        "namespace": session.namespace,
        "idempotency_key": session.idempotency_key,
        "request_hash": session.request_hash,
        "user_id": session.user_id,
        "guest_session_id": session.guest_session_id,
        "session_token": session.session_token,
        "input_type": session.input_type,
        "intent": session.intent,
        "source_channel": session.source_channel,
        "raw_input_text": session.raw_input_text,
        "normalized_text": session.normalized_text,
        "voice_transcript": session.voice_transcript,
        "source_file_name": session.source_file_name,
        "source_file_type": session.source_file_type,
        "source_file_size": session.source_file_size,
        "source_file_path": session.source_file_path,
        "audio_file_name": session.audio_file_name,
        "audio_file_type": session.audio_file_type,
        "audio_file_size": session.audio_file_size,
        "audio_file_path": session.audio_file_path,
        "delivery_location": session.delivery_location,
        "target_currency": session.target_currency,
        "priority": session.priority,
        "status": session.status,
        "parse_status": session.parse_status,
        "analysis_status": session.analysis_status,
        "workflow_status": session.workflow_status,
        "confidence_score": session.confidence_score,
        "warnings": session.warnings or [],
        "suggestions": session.suggestions or [],
        "metadata_json": session.metadata_json or {},
        "parsed_payload": session.parsed_payload or {},
        "normalized_payload": session.normalized_payload or {},
        "analysis_payload": session.analysis_payload or {},
        "preview_payload": session.preview_payload or {},
        "bom_id": session.bom_id,
        "analysis_id": session.analysis_id,
        "project_id": session.project_id,
        "items": [_serialize_item(item) for item in session.items],
        "created_at": session.created_at.isoformat() if session.created_at else None,
        "updated_at": session.updated_at.isoformat() if session.updated_at else None,
    }


def _save_upload_file(upload, subdir: str = "intake") -> Tuple[Optional[str], Optional[int], Optional[str], Optional[str]]:
    if not upload:
        return None, None, None, None

    filename = upload.filename or f"{uuid.uuid4().hex}.bin"
    safe_name = re.sub(r"[^a-zA-Z0-9._-]+", "_", filename)
    target_dir = Path(settings.UPLOAD_DIR) / subdir
    target_dir.mkdir(parents=True, exist_ok=True)
    path = target_dir / f"{uuid.uuid4().hex}_{safe_name}"

    contents = upload.file.read()
    with open(path, "wb") as f:
        f.write(contents)

    return str(path), len(contents), filename, upload.content_type


def parse_intake_payload(
    *,
    db: Session,
    payload: IntakeParseRequest,
    user: Optional[User],
    idempotency_key: Optional[str] = None,
    source_file_name: Optional[str] = None,
    source_file_type: Optional[str] = None,
    source_file_size: Optional[int] = None,
    source_file_path: Optional[str] = None,
    audio_file_name: Optional[str] = None,
    audio_file_type: Optional[str] = None,
    audio_file_size: Optional[int] = None,
    audio_file_path: Optional[str] = None,
    file_text: Optional[str] = None,
) -> Dict[str, Any]:
    raw_text = payload.raw_input_text or ""
    voice_transcript = payload.voice_transcript or ""
    input_type = _detect_input_type(
        explicit_type=payload.input_type.value if hasattr(payload.input_type, "value") else str(payload.input_type),
        raw_input_text=raw_text,
        file_name=source_file_name,
        voice_transcript=voice_transcript,
    )
    intent = _detect_intent(
        explicit_intent=payload.intent.value if hasattr(payload.intent, "value") else str(payload.intent),
        text=" ".join([raw_text or "", voice_transcript or "", file_text or ""]),
    )

    combined_text, parsed_items, summary = _normalize_request_text(
        raw_input_text=raw_text,
        input_type=input_type,
        file_text=file_text,
        voice_transcript=voice_transcript,
    )

    suggestions = _unique_suggestions(
        summary.get("suggestions", [])
        + [SOURCING_FALLBACKS["buyer_action"]]
        + (
            [SOURCING_FALLBACKS["supplier_action"]]
            if intent in {"vendor_search", "deep_search"}
            else []
        )
        + (
            [SOURCING_FALLBACKS["rfq_action"]]
            if intent in {"rfq", "compare"}
            else []
        )
    )

    confidence_score = 0.0
    if parsed_items:
        confidence_score = round(sum(item["confidence"] for item in parsed_items) / len(parsed_items), 2)

    request_hash = _stable_hash(
        {
            "raw_input_text": raw_text,
            "voice_transcript": voice_transcript,
            "file_text": file_text,
            "input_type": input_type,
            "intent": intent,
            "delivery_location": payload.delivery_location,
            "target_currency": payload.target_currency,
            "priority": payload.priority,
            "source_file_name": source_file_name,
            "source_file_type": source_file_type,
            "source_file_size": source_file_size,
            "audio_file_name": audio_file_name,
            "audio_file_type": audio_file_type,
            "audio_file_size": audio_file_size,
        }
    )

    session = _resolve_session_identity(
        db=db,
        namespace="intake.parse",
        idempotency_key=idempotency_key,
        request_hash=request_hash,
        user=user,
        session_token=payload.session_token,
        input_type=input_type,
        intent=intent,
        delivery_location=payload.delivery_location,
        target_currency=payload.target_currency,
        priority=payload.priority,
        raw_input_text=raw_text,
        normalized_text=combined_text,
        voice_transcript=voice_transcript or None,
        source_channel=payload.source_channel,
        metadata=payload.metadata,
        source_file_name=source_file_name,
        source_file_type=source_file_type,
        source_file_size=source_file_size,
        source_file_path=source_file_path,
        audio_file_name=audio_file_name,
        audio_file_type=audio_file_type,
        audio_file_size=audio_file_size,
        audio_file_path=audio_file_path,
    )

    session.input_type = input_type
    session.intent = intent
    session.raw_input_text = raw_text
    session.normalized_text = combined_text
    session.voice_transcript = voice_transcript or None
    session.delivery_location = payload.delivery_location
    session.target_currency = payload.target_currency
    session.priority = payload.priority
    session.confidence_score = confidence_score
    session.warnings = summary.get("warnings", [])
    session.suggestions = suggestions
    session.parsed_payload = {
        "input_type": input_type,
        "intent": intent,
        "parsed_items": parsed_items,
        "parsed_summary": summary,
    }
    session.normalized_payload = {
        "normalized_text": combined_text,
        "parsed_items": parsed_items,
        "parsed_summary": summary,
        "confidence_score": confidence_score,
    }
    session.status = "parsed"
    session.parse_status = "parsed"
    session.analysis_status = "pending"
    session.workflow_status = "parsed"

    db.flush()

    existing_items = db.query(IntakeItem).filter(IntakeItem.session_id == session.id).all()
    for existing in existing_items:
        db.delete(existing)
    db.flush()

    for item in parsed_items:
        db.add(
            IntakeItem(
                session_id=session.id,
                line_no=item["line_no"],
                raw_text=item["raw_text"],
                item_name=item["item_name"],
                category=item["category"],
                material=item["material"],
                process=item["process"],
                quantity=item["quantity"],
                unit=item["unit"],
                specs=item["specs"],
                confidence=item["confidence"],
                warnings=item["warnings"],
                source_payload=item,
            )
        )

    db.flush()

    return {
        "session": session,
        "combined_text": combined_text,
        "parsed_items": parsed_items,
        "summary": summary,
        "confidence_score": confidence_score,
        "input_type": input_type,
        "intent": intent,
        "request_hash": request_hash,
    }


def _build_preview_payload(
    *,
    project,
    analysis,
    strategy: Dict[str, Any],
    procurement: Dict[str, Any],
    session: IntakeSession,
    is_authenticated: bool,
) -> Dict[str, Any]:
    lifecycle = (project.project_metadata or {}).get("analysis_lifecycle", {}) or {}
    workspace_route = lifecycle.get("workspace_route") or f"/project/{project.id}"
    if is_authenticated:
        return {
            "is_preview": False,
            "guest_bom_id": lifecycle.get("guest_bom_id") or str(project.bom_id),
            "session_token": lifecycle.get("session_token") or session.session_token or "",
            "analysis_status": lifecycle.get("analysis_status") or "authenticated_unlocked",
            "report_visibility_level": lifecycle.get("report_visibility_level") or "full",
            "unlock_status": lifecycle.get("unlock_status") or "unlocked",
            "workspace_route": workspace_route,
            "analysis_lifecycle": {
                "guest_bom_id": lifecycle.get("guest_bom_id") or str(project.bom_id),
                "project_id": project.id,
                "session_token": lifecycle.get("session_token") or session.session_token or "",
                "analysis_status": lifecycle.get("analysis_status") or "authenticated_unlocked",
                "report_visibility_level": lifecycle.get("report_visibility_level") or "full",
                "unlock_status": lifecycle.get("unlock_status") or "unlocked",
                "workspace_route": workspace_route,
            },
            "project_id": project.id,
            "bom_id": project.bom_id,
            "analyzer_report": project.analyzer_report or {},
            "strategy": project.strategy or strategy,
            "procurement_plan": project.procurement_plan or procurement,
            "total_parts": project.total_parts,
            "priority": session.priority or "cost",
            "currency": project.currency or strategy.get("currency", session.target_currency or "USD"),
        }

    preview = project_service.build_guest_preview(
        project,
        session_token=session.session_token or "",
        analysis_status="guest_preview",
        report_visibility_level="preview",
        unlock_status="locked",
    )
    preview["workspace_route"] = workspace_route
    return preview


def _build_strategy_and_procurement(
    *,
    db: Session,
    parsed_items: List[Dict[str, Any]],
    delivery_location: str,
    target_currency: str,
    priority: str,
    input_context: Dict[str, Any],
) -> Tuple[Dict[str, Any], Dict[str, Any], Dict[str, Any]]:
    strategy_input = {
        "section_2_component_breakdown": [
            {
                "item_id": f"INTAKE-{item['line_no']}",
                "description": item["item_name"] or item["raw_text"],
                "quantity": item["quantity"],
                "material": item["material"] or "",
                "mpn": "",
                "manufacturer": "",
                "notes": item["raw_text"],
                "category": item["category"],
                "classification_confidence": item["confidence"],
                "geometry": None,
                "tolerance": (item.get("specs") or {}).get("tolerance"),
                "material_form": (item.get("specs") or {}).get("material_hint") or item.get("material"),
                "secondary_ops": [item["process"]] if item["process"] else [],
                "specs": item["specs"],
                "procurement_class": "rfq_required" if item["category"] in {"custom_mechanical", "machined", "sheet_metal", "electronics"} else "catalog_purchase",
                "rfq_required": item["category"] in {"custom_mechanical", "machined", "sheet_metal", "electronics"},
                "drawing_required": item["category"] in {"custom_mechanical", "machined", "sheet_metal"},
                "is_custom": item["category"] in {"custom_mechanical", "machined", "sheet_metal"},
                "part_type": item["category"],
            }
            for item in parsed_items
        ]
    }

    vendor_memories = vendor_service.get_vendor_memories(db)
    strategy = build_strategy_output(
        strategy_input,
        delivery_location,
        vendor_memories,
        pricing_history=[],
        external_pricing=input_context.get("external_pricing", {}),
        db=db,
        priority=priority if priority in ("cost", "speed") else "cost",
        target_currency=target_currency,
    )
    procurement = generate_procurement_plan(
        strategy, target_currency, max_suppliers=5
    )
    return strategy_input, strategy, procurement


def _run_analysis_pipeline(
    *,
    db: Session,
    session: IntakeSession,
    parsed_items: List[Dict[str, Any]],
    combined_text: str,
    summary: Dict[str, Any],
    delivery_location: str,
    target_currency: str,
    priority: str,
    file_bytes: Optional[bytes],
    file_name: Optional[str],
    file_type: Optional[str],
) -> Dict[str, Any]:
    if file_bytes and file_name and file_type and file_name.lower().endswith((".csv", ".xlsx", ".xls", ".tsv", ".txt", ".md", ".pdf")) and session.input_type in {"bom", "file", "auto"}:
        try:
            analyzer_output = analyzer_service.call_analyzer_sync(
                file_bytes=file_bytes,
                filename=file_name,
                user_location=delivery_location,
                target_currency=target_currency,
            )
        except AttributeError:
            # backward compatible fallback
            import asyncio
            analyzer_output = asyncio.get_event_loop().run_until_complete(
                analyzer_service.call_analyzer(
                    file_bytes=file_bytes,
                    filename=file_name,
                    user_location=delivery_location,
                    target_currency=target_currency,
                )
            )
    else:
        analyzer_output = _build_synthetic_analyzer_output(
            session=session,
            parsed_items=parsed_items,
            combined_text=combined_text,
            summary=summary,
            delivery_location=delivery_location,
            target_currency=target_currency,
            priority=priority,
        )

    bom_record = bom_service.create_bom_from_analyzer(
        db,
        analyzer_output,
        file_name=file_name or f"intake-{session.id}.txt",
        file_type=(file_name.rsplit(".", 1)[-1] if file_name and "." in file_name else (file_type or "txt")),
        user_id=session.user_id,
        session_token=session.session_token,
    )

    parts = bom_service.get_bom_parts_as_dicts(db, bom_record.id)
    external_pricing = pricing_service.fetch_external_pricing(parts)
    enriched = pricing_service.enrich_analysis_with_pricing(
        analyzer_output.get("_v2_full_report", analyzer_output),
        db,
        external_pricing,
    )

    strategy_input, strategy, procurement = _build_strategy_and_procurement(
        db=db,
        parsed_items=parsed_items,
        delivery_location=delivery_location,
        target_currency=target_currency,
        priority=priority,
        input_context={"external_pricing": external_pricing},
    )

    ps = strategy.get("procurement_strategy", {})
    cs = ps.get("cost_summary", {})
    rec = strategy.get("recommended_strategy", {})
    cost_range = cs.get("range", [0, 0])

    analysis = AnalysisResult(
        bom_id=bom_record.id,
        user_id=session.user_id,
        guest_session_id=bom_record.guest_session_id,
        raw_analyzer_output=analyzer_output,
        structured_output={
            "strategy": strategy,
            "enriched": {
                "analyzer": enriched,
                "procurement_plan": procurement,
                "external_pricing": {k: v for k, v in external_pricing.items() if v},
                "priority": priority,
                "input_type": session.input_type,
                "intent": session.intent,
            },
        },
        recommended_location=rec.get("location", delivery_location),
        average_cost=cs.get("average", rec.get("average_cost", 0)),
        cost_range_low=cost_range[0] if len(cost_range) > 0 else 0,
        cost_range_high=cost_range[1] if len(cost_range) > 1 else 0,
        savings_percent=cs.get("savings_percent", rec.get("savings_percent", 0)),
        lead_time_days=rec.get("lead_time", 0),
        decision_summary=strategy.get("decision_summary", ""),
        source_version=analyzer_output.get("_meta", {}).get("version", "unknown"),
    )
    db.add(analysis)
    db.flush()

    bom_record.status = "analyzed"

    project = project_service.upsert_project_from_analysis(
        db,
        bom=bom_record,
        analysis=analysis,
        analyzer_output=analyzer_output,
        strategy=strategy,
        procurement=procurement,
    )

    lifecycle = project_service.persist_analysis_lifecycle(
        db,
        bom=bom_record,
        analysis=analysis,
        project=project,
        session_token=session.session_token or bom_record.session_token or "",
        analysis_status="authenticated_unlocked" if session.user_id else "guest_preview",
        report_visibility_level="full" if session.user_id else "preview",
        unlock_status="unlocked" if session.user_id else "locked",
    )

    match_results: list = []
    try:
        source_file = bom_record.source_file_name or file_name or "intake.txt"
        match_results = resolver_service.resolve_and_learn(
            db,
            parts,
            bom_record.id,
            source_file=source_file,
        )
        resolver_service.update_bom_parts_with_matches(
            db,
            bom_record.id,
            match_results,
            parts,
        )
    except Exception as e:
        logger.warning("[Intake Resolver] non-fatal failure: %s", e, exc_info=True)

    try:
        if match_results:
            review_service.create_review_items_from_resolver(
                db, bom_record.id, match_results, parts
            )
    except Exception as e:
        logger.warning("[Intake Review] non-fatal failure: %s", e, exc_info=True)

    session.bom_id = bom_record.id
    session.analysis_id = analysis.id
    session.project_id = project.id
    session.status = "completed"
    session.parse_status = "normalized"
    session.analysis_status = "analyzed"
    session.workflow_status = "completed"
    session.analysis_payload = {
        "analyzer_output": analyzer_output,
        "strategy": strategy,
        "procurement_plan": procurement,
        "external_pricing": external_pricing,
    }
    session.preview_payload = _build_preview_payload(
        project=project,
        analysis=analysis,
        strategy=strategy,
        procurement=procurement,
        session=session,
        is_authenticated=bool(session.user_id),
    )
    db.flush()

    return {
        "session": session,
        "bom": bom_record,
        "analysis": analysis,
        "project": project,
        "lifecycle": lifecycle,
        "strategy": strategy,
        "procurement": procurement,
        "preview": session.preview_payload,
        "normalized_items": parsed_items,
        "summary": summary,
        "analyzer_output": analyzer_output,
    }


def create_or_update_intake(
    *,
    db: Session,
    payload: IntakeParseRequest,
    user: Optional[User],
    idempotency_key: Optional[str],
    file_bytes: Optional[bytes] = None,
    file_name: Optional[str] = None,
    file_type: Optional[str] = None,
    file_path: Optional[str] = None,
    audio_bytes: Optional[bytes] = None,
    audio_name: Optional[str] = None,
    audio_type: Optional[str] = None,
    audio_path: Optional[str] = None,
) -> Dict[str, Any]:
    file_text = None
    if file_bytes and file_name:
        try:
            file_text = _read_file_as_text(file_bytes, file_name)
        except Exception as e:
            logger.warning("Failed to read intake file bytes as text: %s", e, exc_info=True)
            file_text = ""

    parsed = parse_intake_payload(
        db=db,
        payload=payload,
        user=user,
        idempotency_key=idempotency_key,
        source_file_name=file_name,
        source_file_type=file_type,
        source_file_size=len(file_bytes) if file_bytes else None,
        source_file_path=file_path,
        audio_file_name=audio_name,
        audio_file_type=audio_type,
        audio_file_size=len(audio_bytes) if audio_bytes else None,
        audio_file_path=audio_path,
        file_text=file_text,
    )

    session = parsed["session"]
    return {
        "session": session,
        "parsed_items": parsed["parsed_items"],
        "summary": parsed["summary"],
        "confidence_score": parsed["confidence_score"],
        "input_type": parsed["input_type"],
        "intent": parsed["intent"],
        "normalized_text": parsed["combined_text"],
        "file_text": file_text,
        "file_bytes": file_bytes,
        "file_name": file_name,
        "file_type": file_type,
    }


def finalize_intake_submission(
    *,
    db: Session,
    payload: IntakeSubmitRequest,
    user: Optional[User],
    idempotency_key: Optional[str],
    file_bytes: Optional[bytes] = None,
    file_name: Optional[str] = None,
    file_type: Optional[str] = None,
    file_path: Optional[str] = None,
    audio_bytes: Optional[bytes] = None,
    audio_name: Optional[str] = None,
    audio_type: Optional[str] = None,
    audio_path: Optional[str] = None,
    async_finalize: bool = True,
) -> Dict[str, Any]:
    parsed = create_or_update_intake(
        db=db,
        payload=payload,
        user=user,
        idempotency_key=idempotency_key,
        file_bytes=file_bytes,
        file_name=file_name,
        file_type=file_type,
        file_path=file_path,
        audio_bytes=audio_bytes,
        audio_name=audio_name,
        audio_type=audio_type,
        audio_path=audio_path,
    )

    session: IntakeSession = parsed["session"]

    # If user is missing, create a stable guest session token.
    if not session.session_token:
        session.session_token = payload.session_token or uuid.uuid4().hex
        db.flush()

    analysis_payload = _run_analysis_pipeline(
        db=db,
        session=session,
        parsed_items=parsed["parsed_items"],
        combined_text=parsed["normalized_text"],
        summary=parsed["summary"],
        delivery_location=payload.delivery_location,
        target_currency=payload.target_currency,
        priority=payload.priority,
        file_bytes=file_bytes,
        file_name=file_name,
        file_type=file_type,
    )

    session.status = "completed"
    session.parse_status = "normalized"
    session.analysis_status = "analyzed"
    session.workflow_status = "completed"
    db.flush()

    preview = analysis_payload["preview"]
    project = analysis_payload["project"]
    bom_record = analysis_payload["bom"]
    analysis = analysis_payload["analysis"]
    lifecycle = analysis_payload["lifecycle"]
    strategy = analysis_payload["strategy"]
    procurement = analysis_payload["procurement"]

    return {
        "session": session,
        "bom_id": bom_record.id,
        "project_id": project.id,
        "analysis_id": analysis.id,
        "workspace_route": f"/project/{project.id}",
        "analysis_status": lifecycle.get("analysis_status", "analyzed"),
        "report_visibility_level": lifecycle.get("report_visibility_level", "preview"),
        "unlock_status": lifecycle.get("unlock_status", "locked"),
        "normalized_items": parsed["parsed_items"],
        "analysis_lifecycle": lifecycle,
        "preview": preview,
        "strategy": strategy,
        "procurement_plan": procurement,
        "parsed_summary": parsed["summary"],
        "warnings": session.warnings or [],
        "suggestions": session.suggestions or [],
    }


def list_sessions(
    db: Session,
    *,
    user: Optional[User] = None,
    session_token: Optional[str] = None,
    limit: int = 20,
    offset: int = 0,
) -> Tuple[List[IntakeSession], int]:
    q = db.query(IntakeSession)
    if user:
        q = q.filter(IntakeSession.user_id == user.id)
    elif session_token:
        q = q.filter(IntakeSession.session_token == session_token)
    else:
        q = q.filter(func.false())  # type: ignore[attr-defined]

    total = q.count()
    items = q.order_by(IntakeSession.created_at.desc()).offset(offset).limit(limit).all()
    return items, total


def get_session(db: Session, session_id: str) -> Optional[IntakeSession]:
    return db.query(IntakeSession).filter(IntakeSession.id == session_id).first()