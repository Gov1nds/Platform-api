"""Excel report exporter (Blueprint §15, C27)."""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def render_xlsx(report_type: str, data: dict) -> bytes:
    wb = Workbook(); ws = wb.active
    ws.title = report_type.replace("_", " ").title()[:31]
    headers = data.get("columns", [])
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for row in data.get("rows", []):
        ws.append(row)
    for sub in data.get("subsheets", []):
        ws2 = wb.create_sheet(sub["name"][:31])
        ws2.append(sub.get("columns", []))
        for r in sub.get("rows", []):
            ws2.append(r)
    buf = BytesIO(); wb.save(buf); return buf.getvalue()
